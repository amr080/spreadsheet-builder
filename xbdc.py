import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment

bdc_data_path = 'data/bdc.xlsx'
bdc_data = pd.read_excel(bdc_data_path, engine='openpyxl')
column_mappings = {
    'CUSIP': 'LN',
    'Ticker': 'B',
    'Shares Outstanding': 'BR',
    'Market Price': 'E',
    'Distribution Amount': 'LW',
    'Interest Income': 'AA',
    'Dividend Income': 'AB',
    'G&A Fee %': 'IX',
    'Incentive Fee': 'AJ',
    'CG Incentive Fee': 'LJ',
    'NII Incentive Fee': 'LG',
    'Stated Base Mgmt Fee (%)': 'LE',
    'Base Management Fee %': 'IZ',
    'Div Freq': 'L',
    'NII': 'X',
    'Fund Name': 'C',
    'CIK': 'KE',
    'Market Cap': 'AH',
    'NAV': 'F',
    'Net Assets': 'JN',
    'Total Assets': 'LO',
    'Inception Assets (millions)': 'BE',
    'Inception Price ($)': 'BD',
    'ROE Inception': 'LV',
    'Inception Date': 'BB',
    'Core NII': 'BN',
    'NII / Share': 'Y',
    'UNII/Share': 'BP',
    'Expense Ratio': 'AI',
    'Total Investments': 'AR',
    'Employees': 'FH',
    'YTD Price TR': 'AW',
    'Listed': 'LL'
}


def column_index(column_letter):
    index = 0
    for char in column_letter:
        index = index * 26 + (ord(char.upper()) - ord('A') + 1)
    return index - 1

selected_columns = {key: bdc_data.iloc[:, column_index(value)] for key, value in column_mappings.items()}
selected_data = pd.DataFrame(selected_columns)

timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
output_filename = f"xbdc_{timestamp}.xlsx"
output_path = f"output/{output_filename}"

# Save the DataFrame to an Excel file
selected_data.to_excel(output_path, index=False)

# Load the workbook
wb = load_workbook(output_path)
ws = wb.active

# Apply formatting to the first row
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.border = Border(bottom=Side(style='thin'))
    cell.alignment = Alignment(horizontal='left')

# Save the workbook
wb.save(output_path)

print(f"Data saved to {output_path}")