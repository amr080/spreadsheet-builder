import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment

cef_data_path = 'data/cef.xlsx'
cef_data = pd.read_excel(cef_data_path, engine='openpyxl')

column_mappings = {
    'Ticker': 'B',
    'CUSIP': 'SX',
    'Net Assets': 'BR',
    'Distrib Amount': 'Y',
    'Frequency': 'AC',
    'Total Expenses': 'RK',
    'Mgmt Fee': 'RG',
    'Interest Expense': 'RH',
    'Preferred Expense': 'RI',
    'Other Expense': 'RJ',
    'Director / Trustee Compensation': 'LB',
    'Net Expenses': 'RL',
    'UNII': 'AO',
    'UNII Freq': 'PL',
    'Gross Assets': 'HJ',
    'Gross Expense': 'HK',
    'Gross vs Net Assets': 'MM',
    'Name': 'D',
    'CIK': 'QE',
    'Shares Outstanding': 'BS',
    'Market Cap': 'PG',
    'Market Price': 'G',
    'NAV': 'H',
    'Fair Market Value': 'LC',
    'Exp Ratio': 'BP',
    'Realized Cap Gain': 'PI',
    'Listed': 'SV'
}

def column_index(column_letter):
    index = 0
    for char in column_letter:
        index = index * 26 + (ord(char.upper()) - ord('A') + 1)
    return index - 1

selected_columns = {key: cef_data.iloc[:, column_index(value)] for key, value in column_mappings.items()}
selected_data = pd.DataFrame(selected_columns)

timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
output_filename = f"xcef_{timestamp}.xlsx"
output_path = f"output/{output_filename}"

# Save the DataFrame to an Excel file
selected_data.to_excel(output_path, index=False)

# Load the workbook
wb = load_workbook(output_path)
ws = wb.active

# Apply formatting to the first row
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.border = Border(bottom=Side(style='thin'))
    cell.alignment = Alignment(horizontal='left')

# Save the workbook
wb.save(output_path)

print(f"Data saved to {output_path}")